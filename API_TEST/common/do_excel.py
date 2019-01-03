# _*_coding:utf-8_*_
"""
@author:malieyong
@time:2018/12/18 22:25
"""


from common.config import Conf
from common import request
import openpyxl
import json
import os

class Case:
    def __init__(self):
        self.id = None
        self.api = None
        self.api_describe = None
        self.api_name = None
        self.method = None
        self.url = None
        self.request_data = None
        self.expected_data = None
        self.practical_data = None

class Do_Excel:

    # 打开指定文件
    def __init__(self,file_name):
        try:
            self.file_name = file_name
            self.workbook = openpyxl.load_workbook(filename=file_name)
        except FileNotFoundError as e:
            print('{0} not found ,please check  path '.format(file_name))
            raise e

    # 获取指定表单数据
    def get_cases(self,sheet_name):
        sheet = self.workbook[sheet_name]
        cases = []
        for row in range(2,sheet.max_row+1):
            case = Case()
            case.id = sheet.cell(row,1).value             #序号
            case.api = sheet.cell(row,2).value            #api
            case.api_describe = sheet.cell(row,3).value   #api描述
            case.api_name = sheet.cell(row,4).value       #api名字
            case.method = sheet.cell(row,5).value         #请求方式
            case.url = sheet.cell(row,6).value            #请求地址
            case.request_data = sheet.cell(row,7).value   #请求数据
            case.expect = sheet.cell(row,8).value         #期望结果
            case.practical =sheet.cell(row,9).value       #实际结果
            case.result = sheet.cell(row,10).value        #测试结果
            cases.append(case)
        return cases

    # 获取所有表单名字
    def get_all_sheet_names(self):
        sheet_names = self.workbook.get_sheet_names()
        # print(sheet_names)
        return sheet_names

    #回写测试用例返回数据practical 以及测试结果result
    def write_practical_by_case_id(self,sheet_name,case_id,practical,result):
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        for r in range(2,max_row+1):
            if sheet.cell(r,1).value == case_id:
                sheet.cell(r,9).value = practical
                sheet.cell(r,10).value = result
                self.workbook.save(self.file_name)
                break

    #回写替换后的请求数据
    def write_request_by_case_id(self,sheet_name,case_id,replace_end):
        sheet = self.workbook[sheet_name]
        max_row = sheet.max_row
        for r in range(2, max_row + 1):
            #判断cese_id 若值相等，进行回写相应列
            if sheet.cell(r, 1).value == case_id:
                sheet.cell(r, 7).value = replace_end
                self.workbook.save(self.file_name)
                break


    #拼接url地址
    def join_url(self,sheet_url):
        #获取基础路径中url
        conf = Conf()
        base_url = conf.get_base_url('api','url')
        # print(type(base_url))
        return os.path.join(base_url,sheet_url)




if __name__ == '__main__':
    do_excel = Do_Excel('../datas./testdatas.xlsx')
    datas = do_excel.get_cases('login')
    for item in datas:
        item_url = do_excel.join_url(sheet_url=item.url)
        print(item_url)
        python_str= json.loads(item.request_data)
        try:
            res = request.Request(url=item_url,method=item.method,data=python_str)
            print(res.get_josn())
        except SyntaxError as e:
            print('第{0}用例执行失败,失败名称{1}'.format(item.id,item.api_describe))
